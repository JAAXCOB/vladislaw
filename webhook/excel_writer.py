"""
Phase 3 — write extracted job data into the existing Excel file.

Finds the correct monthly sheet by message date, appends one row:
  Дата | VIN/Гос.номер ТС | Услуга | Сумма
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from webhook.schema import ExtractedJob

log = logging.getLogger("max_webhook.excel")

MOSCOW_TZ = timezone(timedelta(hours=3))

# Matches sheet names in the real file (note: "Сентябро" is a typo in the original)
MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябро",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

# Matches the existing sheet's cell style (Calibri 11, centered, mm-dd-yy dates)
DEFAULT_FONT = Font(name="Calibri", size=11)
CENTER_ALIGN = Alignment(horizontal="center")
DATE_FORMAT = "mm-dd-yy"

# Light fill for rows that need human re-checking (needs_review=True)
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _first_empty_row(ws: Worksheet, key_column: int = 1) -> int:
    """
    Returns the first row (below the header) whose key_column cell is empty.

    ws.max_row / ws.append() can't be trusted here: openpyxl counts a row
    as "used" if it ever had a value or style applied, even after the
    value was manually cleared in Excel. That leaves stale formatting
    far below the real data and makes append() start writing hundreds of
    rows past the actual empty area. Scanning for a genuinely empty cell
    finds the real gap instead.
    """
    row = 2
    while ws.cell(row=row, column=key_column).value is not None:
        row += 1
    return row


def _sheet_name(dt: datetime) -> str:
    """Return sheet name for a given date, e.g. 'Август 26'."""
    month = MONTH_NAMES[dt.month]
    year = str(dt.year)[2:]
    return f"{month} {year}"



def _format_services(job: ExtractedJob) -> str:
    """
    Format services list to match the existing sheet style, e.g.
    'Эвакуация + 25 км за МКАД' or 'Ложная подача'.
    No prices in this column — the total goes in the Сумма column.
    """
    if not job.services:
        return ""
    parts = [svc.name.strip().capitalize() for svc in job.services if svc.name.strip()]
    return " + ".join(parts)


def append_job(
    excel_path: str | Path,
    job: ExtractedJob,
    message_timestamp_ms: int,
    original_text: str = "",
) -> str:
    """
    Append one row to the appropriate monthly sheet, chosen automatically
    from the message date (e.g. a message on Sep 1 goes to 'Сентябро 26'
    even if the previous message was in 'Август 26').

    Returns the sheet name where the row was written.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    dt = datetime.fromtimestamp(message_timestamp_ms / 1000, tz=MOSCOW_TZ)
    sheet_name = _sheet_name(dt)

    wb = openpyxl.load_workbook(path)

    if sheet_name not in wb.sheetnames:
        log.warning("Sheet '%s' not found in %s — available: %s", sheet_name, path.name, wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")

    ws = wb[sheet_name]

    plate = job.license_plate or f"[НЕТ НОМЕРА] {original_text[:30]}"
    service_text = _format_services(job) or original_text[:60]
    amount = job.total_amount_rub

    target_row = _first_empty_row(ws)
    row = [dt.date(), plate, service_text, amount]
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=target_row, column=col_idx).value = value

    # Apply the same style as existing rows: Calibri 11, centered, mm-dd-yy dates
    for col in range(1, 5):
        cell = ws.cell(row=target_row, column=col)
        cell.font = DEFAULT_FONT
        cell.alignment = CENTER_ALIGN
        if job.needs_review:
            cell.fill = REVIEW_FILL

    ws.cell(row=target_row, column=1).number_format = DATE_FORMAT

    wb.save(path)

    log.info(
        "EXCEL | sheet='%s' | row=%d | plate=%s | amount=%s | review=%s",
        sheet_name,
        target_row,
        plate,
        amount,
        job.needs_review,
    )

    return sheet_name
