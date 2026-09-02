"""
Writes each closed job into the payroll Excel file, placing the amount
into the correct employee's column.

Sheet layout observed in the real file:
  Дата | VIN/Гос.номер ТС | Услуга | <Сотрудник 1> | <Сотрудник 2> | ... | 750

Only ONE employee column is filled per row (the one who did the job);
all others stay blank. The last column's header is a plain number (750)
rather than a name — per explicit instruction, this column is never
touched by automation. To stay robust if more such columns are added
later, any header that isn't a string is treated as protected and
excluded from employee matching.
"""
from __future__ import annotations

import logging
from copy import copy
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from webhook.excel_writer import _first_empty_row, _format_services
from webhook.schema import ExtractedJob

log = logging.getLogger("max_webhook.payroll")

MOSCOW_TZ = timezone(timedelta(hours=3))

# Each month has one shared payroll sheet. Employees are columns inside that
# sheet — never separate worksheets.
MONTH_NAMES = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

FIXED_COLUMNS = {"Дата", "VIN/Гос.номер ТС", "Услуга"}

DEFAULT_FONT = Font(name="Calibri", size=11)
CENTER_ALIGN = Alignment(horizontal="center")
DATE_FORMAT = "mm-dd-yy"

# Light fill for rows where the employee couldn't be confidently matched
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


class AmbiguousEmployeeError(Exception):
    """Raised when the sender name matches zero or multiple employee columns."""


def _sheet_name(dt: datetime) -> str:
    return MONTH_NAMES[dt.month]


def _get_or_create_sheet(wb: openpyxl.Workbook, sheet_name: str, path: Path) -> Worksheet:
    """
    Return one shared sheet for the requested month. If September does not
    exist yet, create exactly one "Сентябрь" sheet using August as its
    structure: same employee columns, header styles and column widths.
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    month_number = next(number for number, name in MONTH_NAMES.items() if name == sheet_name)
    if month_number == 1:
        raise ValueError(f"Cannot create January payroll sheet automatically in {path.name}")

    previous_month_name = MONTH_NAMES[month_number - 1]
    if previous_month_name not in wb.sheetnames:
        raise ValueError(
            f"Previous payroll sheet '{previous_month_name}' not found in {path.name}"
        )

    template_ws = wb[previous_month_name]
    ws = wb.create_sheet(sheet_name)

    for col_idx in range(1, template_ws.max_column + 1):
        src = template_ws.cell(row=1, column=col_idx)
        dst = ws.cell(row=1, column=col_idx)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.alignment = copy(src.alignment)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.number_format = src.number_format

    for col_letter, col_dim in template_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = col_dim.width

    log.info(
        "Created shared payroll sheet '%s' from '%s' in %s",
        sheet_name,
        previous_month_name,
        path.name,
    )
    return ws


def _employee_columns(ws: Worksheet) -> dict[int, str]:
    """
    Maps column index -> employee display name, for header cells that are
    text (skips Дата/VIN/Услуга and any numeric-header column like 750).
    """
    columns: dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if isinstance(header, str) and header.strip() and header.strip() not in FIXED_COLUMNS:
            columns[col_idx] = header.strip()
    return columns


def _match_employee_column(ws: Worksheet, employee_name: str) -> Optional[int]:
    """
    Finds the single column whose header matches employee_name.
    Returns None if there is no match or more than one — callers must
    never guess when it's ambiguous, this is payroll data.
    """
    if not employee_name:
        return None

    name_lower = employee_name.strip().lower()
    candidates = _employee_columns(ws)

    # Pass 1: exact word match (e.g. sender "Валера" vs header "Баранов Валерий Валера")
    exact_matches = [
        col for col, header in candidates.items()
        if name_lower in [w.lower() for w in header.split()]
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]

    # Pass 2: substring fallback, only if still unambiguous
    if not exact_matches:
        substring_matches = [
            col for col, header in candidates.items()
            if name_lower in header.lower() or header.lower() in name_lower
        ]
        if len(substring_matches) == 1:
            return substring_matches[0]

    return None


def append_salary_row(
    payroll_path: str | Path,
    job: ExtractedJob,
    message_timestamp_ms: int,
    employee_name: str,
    original_text: str = "",
) -> tuple[str, bool]:
    """
    Appends one row to the shared monthly payroll sheet: Дата/VIN/Услуга
    are always filled, the amount is placed under the matched employee's
    column only if the match is unambiguous. All employees stay on the same
    worksheet as separate columns.

    Returns (sheet_name, matched) — matched=False means the row was
    written but no employee column could be confidently identified, so
    the amount cell was left blank for manual entry.
    """
    path = Path(payroll_path)
    if not path.exists():
        raise FileNotFoundError(f"Payroll file not found: {path}")

    dt = datetime.fromtimestamp(message_timestamp_ms / 1000, tz=MOSCOW_TZ)
    sheet_name = _sheet_name(dt)

    wb = openpyxl.load_workbook(path)
    ws = _get_or_create_sheet(wb, sheet_name, path)

    plate = job.license_plate or f"[НЕТ НОМЕРА] {original_text[:30]}"
    service_text = _format_services(job) or original_text[:60]
    amount = job.total_amount_rub

    matched_col = _match_employee_column(ws, employee_name) if amount is not None else None

    row_values = [None] * ws.max_column
    row_values[0] = dt.date()
    row_values[1] = plate
    row_values[2] = service_text
    if matched_col is not None:
        row_values[matched_col - 1] = amount

    target_row = _first_empty_row(ws)
    for col_idx, value in enumerate(row_values, 1):
        ws.cell(row=target_row, column=col_idx).value = value

    matched = matched_col is not None

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=target_row, column=col)
        cell.font = DEFAULT_FONT
        cell.alignment = CENTER_ALIGN
        if not matched:
            cell.fill = REVIEW_FILL
    ws.cell(row=target_row, column=1).number_format = DATE_FORMAT

    wb.save(path)

    log.info(
        "PAYROLL | sheet='%s' | row=%d | plate=%s | employee=%s | amount=%s | matched=%s",
        sheet_name, target_row, plate, employee_name, amount, matched,
    )

    return sheet_name, matched
